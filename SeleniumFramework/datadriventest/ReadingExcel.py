from base.BasePage import BaseClass
import openpyxl
import pandas as pd
from datadriventest import XLutils
from xlrd import open_workbook
import xlrd
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementNotVisibleException, NoSuchElementException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec


class ReadingFile(BaseClass):

    def ReadFile(self):
        # File-->Workbook--->Sheets-->Rows-->Cells
        # file = "C:/Users/lukel/Desktop/appium_test/Selenium_Data.xlsx"
        # workbook = openpyxl.load_workbook(file)
        # sheet = workbook["Name"]

        # rows = sheet.max_row  # count number of rows in a excel sheet 6
        # cols = sheet.max_column  # count number of columns in a excel sheet  4

        #print(data_table)

        # # Reading all the rows & columns from excel sheet
        # for r in range(2, rows + 1):
        #     for c in range(1, cols + 1):
        #         #x = d.append(sheet.cell(r, c).value)
        #         print(sheet.cell(r, c).value, end='    ')
        #     print()

        #df = pd.read_excel('C:/Users/lukel/Desktop/appium_test/Selenium_Data.xlsx')
        # List_order = df['Name'].tolist()
        #List_order2 = df['Salary'].tolist()
        # print(List_order)

        # print(List_order2)
        # for i in Selenium_Data.xlsx.index:
        #     order = df.at[i, '']
        #     List_order.append(order)

       # workbook = xlrd.open_workbook('C:/Users/lukel/Desktop/appium_test/Selenium_Data.xlsx')

        file = "C:/Users/lukel/Desktop/appium_test/Selenium_Data.xlsx"
        rows = XLutils.getRowCount(file, "Name")

        # looping and reading data from excel
        for r in range(2, rows + 1):
            # reading data from excel
            name = XLutils.readData(file, "Name", r, 1)
            country = XLutils.readData(file, "Name", r, 2)
            cty = XLutils.readData(file, "Name", r, 3)
            salary = XLutils.readData(file, "Name", r, 4)

        # # get data from table application to verify with data from spreadsheet.
        # wait = WebDriverWait(self.driver, 25, poll_frequency=1,
        #                      ignored_exceptions=[ElementNotVisibleException, NoSuchElementException])
        #
        # data_table = [x.text for x in
        #               wait.until(ec.presence_of_all_elements_located((By.XPATH, "//*[@id='tt']/div/table/tbody/tr")))]
        # print(data_table)
